import asyncio
import logging

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from sdp_lib import logging_config
from sdp_lib.data_capture.events import Event
from sdp_lib.data_capture.main_loop import main
from sdp_lib.data_capture.producers import StageProducer
from sdp_lib.data_capture.storage import RecordsStorage, LogWriter
from sdp_lib.management_controllers.api import PotokS
from sdp_lib.modbus.client import AsyncModbus, create_bits_pattern
from sdp_lib.modbus.data_helpers import Description
from sdp_lib.modbus.fields import FieldNames


def load_or_make_wb(filenname):
    try:
        wb = load_workbook(filenname)
    except FileNotFoundError:
        wb = Workbook()
        wb.active.append(['Время', 'Протокол', 'Номер Фазы', 'Продолжительность фазы'])
        wb.save(filenname)
    return wb


if __name__ == '__main__':
    _wb_filename = 'log_mb_snmp.xlsx'
    _wb = load_or_make_wb(_wb_filename)
    logger = logging.getLogger('penetrate_stage_log')
    log_records_storage = RecordsStorage()
    _log_writer = LogWriter(log_records_storage, logger, _wb, _wb_filename)
    host_snmp = PotokS(ipv4='91.227.113.186', host_id='laba_test')

    stage1 = Description('(Stage=1)', 1)
    stage4 = Description('(Stage=4)', 4)
    stage5 = Description('(Stage=5)', 5)
    stage7 = Description('(Stage=7)', 7)
    bit_states_to_description = {
        create_bits_pattern([0, 0, 0, 0, 1, 0, 1, 0]):                               stage1,
        create_bits_pattern([False, False, True, False, True, False, False, False]): stage4,
        create_bits_pattern([0, 0, 0, 1, 1, 0, 0, 0]):                               stage5,
        create_bits_pattern('00001100'):                                             stage7,
    }
    bit_address_to_description = {6: stage1, 2: stage4, 3: stage5, 5:stage7}
    host_modbus = AsyncModbus(
        ipv4='91.227.113.186',
        port=502,
        matched_bit_states_to_description=bit_states_to_description,
        matched_bit_addr_to_description=bit_address_to_description
    )
    _producers = [
        StageProducer(
            host_snmp,
            Event('Фаза snmp', log_records_storage, protocol=host_snmp.protocol),
            name='snmp_change_stage'
        ),
        StageProducer(
            host_modbus,
            Event('Фаза modbus', log_records_storage, protocol=host_modbus.protocol),
            name='modbus_change_stage'
        )
    ]

    asyncio.run(main(.2, _producers, _log_writer))